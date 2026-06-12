import json
import mimetypes
import os
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from common.config import AppSettings
from common.db import CatalogTemplateItem, Entry, EntryCatalogItem, EntryItemImage, OrgUnit, get_session
from common.repositories.entry_repo import get_default_template_id
from common.services.crypto_service import CryptoService, encrypt_image


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tif", ".tiff"}


def _s(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm(value: Any) -> str:
    return re.sub(r"\s+", "", _s(value)).lower()


def _natural_key(path: str):
    name = os.path.basename(path)
    return [int(p) if p.isdigit() else p.lower() for p in re.split(r"(\d+)", name)]


def _find_excel_file(folder: str) -> Optional[str]:
    preferred = os.path.join(folder, "干部基本信息和档案目录.xlsx")
    if os.path.exists(preferred):
        return preferred
    files = []
    for name in os.listdir(folder):
        path = os.path.join(folder, name)
        if os.path.isfile(path) and name.lower().endswith((".xlsx", ".xlsm")) and not name.startswith("~$"):
            files.append(path)
    files.sort(key=_natural_key)
    return files[0] if files else None


def _folder_name_info(folder: str) -> Dict[str, str]:
    base = os.path.basename(os.path.normpath(folder))
    m = re.search(r"(\d{17}[\dXx])", base)
    id_card = m.group(1).upper() if m else ""
    name = base
    if id_card:
        name = base.replace(m.group(1), "")
    name = name.replace("人事档案", "").replace("干部档案", "").strip(" -_（）()")
    return {"name": name, "id_card": id_card}


def _parse_basic_info(wb, folder: str) -> Dict[str, Any]:
    ws = wb["基本信息"] if "基本信息" in wb.sheetnames else wb.worksheets[0]
    header_row = None
    headers: Dict[str, int] = {}
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [_s(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "姓名" in vals and ("身份证号" in vals or "工号" in vals):
            header_row = r
            headers = {v: i + 1 for i, v in enumerate(vals) if v}
            break
    data: Dict[str, Any] = {"custom_fields": []}
    if header_row is not None:
        row = None
        for r in range(header_row + 1, ws.max_row + 1):
            vals = [_s(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            if any(vals):
                row = r
                break
        if row is not None:
            def get(*names: str) -> str:
                for name in names:
                    col = headers.get(name)
                    if col:
                        return _s(ws.cell(row, col).value)
                return ""
            data.update({
                "name": get("姓名"),
                "emp_no": get("工号", "编号"),
                "role_title": get("岗位", "职务", "职务职级"),
                "phone": get("电话", "联系电话"),
                "id_card": get("身份证号", "身份证", "身份证号码"),
                "status": get("状态"),
            })
            fixed = {"序号", "姓名", "工号", "编号", "岗位", "职务", "职务职级", "电话", "联系电话", "身份证号", "身份证", "身份证号码", "状态", "所属机构"}
            for h, col in headers.items():
                if h in fixed:
                    continue
                value = _s(ws.cell(row, col).value)
                if value:
                    data["custom_fields"].append({"field_name": h, "field_value": value})
    fallback = _folder_name_info(folder)
    if not data.get("name"):
        data["name"] = fallback.get("name", "")
    if not data.get("id_card"):
        data["id_card"] = fallback.get("id_card", "")
    return data


def _parse_catalog_rows(wb) -> List[Dict[str, Any]]:
    if "目录" in wb.sheetnames:
        ws = wb["目录"]
    elif len(wb.worksheets) > 1:
        ws = wb.worksheets[1]
    else:
        ws = wb.worksheets[0]
    rows: List[Dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        vals = [_s(ws.cell(r, c).value) for c in range(1, 8)]
        joined = "".join(vals)
        if not joined:
            continue
        if "干部人事档案目录" in joined or "材料形成时间" in joined:
            continue
        if vals[0] == "序号" or vals[1] == "材料名称" or joined in {"年月日", "序号材料名称年月日页数备注"}:
            continue
        row = {
            "serial": vals[0],
            "name": vals[1],
            "year": vals[2],
            "month": vals[3],
            "day": vals[4],
            "pages": _to_int(vals[5]),
            "remark": vals[6],
        }
        row["is_heading"] = _is_heading_row(row)
        if any(_s(row.get(k)) for k in ("serial", "name", "year", "month", "day", "remark")) or row.get("pages"):
            rows.append(row)
    return rows


def _to_int(value: Any) -> Optional[int]:
    text = _s(value)
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        m = re.search(r"\d+", text)
        return int(m.group(0)) if m else None


def _serial_text(value: Any) -> str:
    return _s(value).replace("－", "-").replace("—", "-").replace("．", ".").strip(" .、")


def _is_chinese_heading_serial(value: Any) -> bool:
    text = _serial_text(value)
    return bool(text and re.fullmatch(r"[一二三四五六七八九十百]+", text))


def _is_sub_heading_serial(value: Any) -> bool:
    text = _serial_text(value)
    return bool(text and re.fullmatch(r"\d+[-.]\d+(?:[-.]\d+)?", text))


def _is_plain_item_serial(value: Any) -> bool:
    text = _serial_text(value)
    return bool(text and re.fullmatch(r"\d+", text))


def _is_heading_row(row: Dict[str, Any]) -> bool:
    serial = row.get("serial")
    name = _s(row.get("name"))
    if not name:
        return False
    if _is_chinese_heading_serial(serial) or _is_sub_heading_serial(serial):
        return True
    if _is_plain_item_serial(serial):
        return False
    if serial and name.endswith("材料") and not (_s(row.get("year")) or _s(row.get("month")) or _s(row.get("day"))):
        return True
    return False


def _org_path(session, org_unit_id: Optional[int]) -> str:
    if org_unit_id is None:
        return ""
    rows = session.query(OrgUnit.id, OrgUnit.parent_id, OrgUnit.name).all()
    by_id = {int(r.id): r for r in rows}
    current = int(org_unit_id)
    parts: List[str] = []
    seen = set()
    while current is not None and current not in seen:
        seen.add(current)
        row = by_id.get(current)
        if not row:
            break
        name = _s(row.name)
        if name:
            parts.append(name)
        current = int(row.parent_id) if row.parent_id is not None else None
    return "/".join(reversed(parts))


def _merge_custom_fields(old_raw: str, imported: List[Dict[str, str]]) -> str:
    merged: Dict[str, str] = {}
    if old_raw:
        try:
            old_items = json.loads(old_raw)
        except Exception:
            old_items = []
        if isinstance(old_items, list):
            for item in old_items:
                if isinstance(item, dict):
                    name = _s(item.get("field_name"))
                    value = _s(item.get("field_value"))
                    if name and value:
                        merged[name] = value
    for item in imported or []:
        name = _s(item.get("field_name"))
        value = _s(item.get("field_value"))
        if name and value:
            merged[name] = value
    if not merged:
        return ""
    return json.dumps([{"field_name": k, "field_value": v} for k, v in merged.items()], ensure_ascii=False)


def _has_catalog_content(session, entry_id: int) -> bool:
    return session.query(EntryCatalogItem.id).filter(EntryCatalogItem.entry_id == int(entry_id)).first() is not None


def _upsert_entry(session, basic: Dict[str, Any], template_id: int, org_unit_id: Optional[int]) -> Tuple[int, bool, str]:
    name = _s(basic.get("name"))
    emp_no = _s(basic.get("emp_no"))
    id_card = _s(basic.get("id_card")).upper()
    entry = None
    match_key = "new"
    if id_card:
        entry = session.query(Entry).filter(Entry.id_card == id_card).order_by(Entry.id.asc()).first()
        if entry is not None:
            match_key = "id_card"
    if entry is None and emp_no:
        entry = session.query(Entry).filter(Entry.emp_no == emp_no, Entry.template_id == int(template_id)).order_by(Entry.id.asc()).first()
        if entry is not None:
            match_key = "emp_no"
    if entry is None and name:
        q = session.query(Entry).filter(Entry.name == name, Entry.template_id == int(template_id))
        if org_unit_id is None:
            q = q.filter(Entry.org_unit_id.is_(None))
        else:
            q = q.filter(Entry.org_unit_id == int(org_unit_id))
        matches = q.order_by(Entry.id.asc()).limit(2).all()
        if len(matches) == 1 and not _has_catalog_content(session, int(matches[0].id)):
            entry = matches[0]
            match_key = "name_empty"
    created = entry is None
    path = _org_path(session, org_unit_id)
    custom_fields = _merge_custom_fields(getattr(entry, "custom_fields", "") if entry else "", basic.get("custom_fields") or [])
    if entry is None:
        entry = Entry(owner_id=1, template_id=int(template_id))
        session.add(entry)
        session.flush()
    if name:
        entry.name = name
    if emp_no:
        entry.emp_no = emp_no
    if _s(basic.get("role_title")):
        entry.role_title = _s(basic.get("role_title"))
    if _s(basic.get("phone")):
        entry.phone = _s(basic.get("phone"))
    if _s(basic.get("status")):
        entry.status = _s(basic.get("status"))
    if id_card:
        entry.id_card = id_card
    if custom_fields:
        entry.custom_fields = custom_fields
    entry.template_id = int(template_id)
    entry.org_unit_id = int(org_unit_id) if org_unit_id is not None else None
    entry.org_path = path
    session.flush()
    return int(entry.id), created, match_key


def _template_nodes(session, template_id: int) -> List[Dict[str, Any]]:
    rows = session.query(CatalogTemplateItem).filter(CatalogTemplateItem.template_id == int(template_id)).order_by(CatalogTemplateItem.sort_order.asc(), CatalogTemplateItem.id.asc()).all()
    by_parent: Dict[Optional[int], List[CatalogTemplateItem]] = {}
    for row in rows:
        by_parent.setdefault(int(row.parent_id) if row.parent_id is not None else None, []).append(row)
    out: List[Dict[str, Any]] = []
    def walk(parent_id: Optional[int]):
        for row in by_parent.get(parent_id, []) or []:
            node = {
                "id": int(row.id),
                "parent_id": int(row.parent_id) if row.parent_id is not None else None,
                "serial": row.serial or "",
                "name": row.name or "",
                "structural": bool(_s(row.serial) or _s(row.name)),
            }
            out.append(node)
            walk(int(row.id))
    walk(None)
    return out


def _create_template_item(session, template_id: int, parent_id: Optional[int], serial: str = "", name: str = "") -> int:
    q = session.query(CatalogTemplateItem.sort_order).filter(CatalogTemplateItem.template_id == int(template_id))
    if parent_id is None:
        q = q.filter(CatalogTemplateItem.parent_id.is_(None))
    else:
        q = q.filter(CatalogTemplateItem.parent_id == int(parent_id))
    row = q.order_by(CatalogTemplateItem.sort_order.desc()).first()
    sort_order = int(row[0] or 0) + 1 if row else 1
    obj = CatalogTemplateItem(template_id=int(template_id), parent_id=int(parent_id) if parent_id is not None else None, sort_order=sort_order, serial=serial or None, name=name or None)
    session.add(obj)
    session.flush()
    return int(obj.id)


def _ec_score(obj: EntryCatalogItem) -> int:
    score = 0
    for attr in ("serial", "name", "year", "month", "day", "remark", "attachment_path"):
        if _s(getattr(obj, attr, "")):
            score += 1
    if obj.pages is not None:
        score += 1
    return score


def _ensure_ec(session, entry_id: int, template_item_id: int) -> EntryCatalogItem:
    rows = session.query(EntryCatalogItem).filter(EntryCatalogItem.entry_id == int(entry_id), EntryCatalogItem.template_item_id == int(template_item_id)).order_by(EntryCatalogItem.id.asc()).all()
    if rows:
        return max(rows, key=lambda o: (_ec_score(o), -int(o.id)))
    obj = EntryCatalogItem(entry_id=int(entry_id), template_item_id=int(template_item_id))
    session.add(obj)
    session.flush()
    return obj


def _reset_entry_catalog_for_import(session, entry_id: int) -> None:
    ec_ids = [
        int(row[0])
        for row in session.query(EntryCatalogItem.id).filter(EntryCatalogItem.entry_id == int(entry_id)).all()
    ]
    if not ec_ids:
        return
    session.query(EntryItemImage).filter(
        EntryItemImage.entry_catalog_item_id.in_(ec_ids)
    ).update({EntryItemImage.original_id: None}, synchronize_session=False)
    session.query(EntryItemImage).filter(
        EntryItemImage.entry_catalog_item_id.in_(ec_ids)
    ).delete(synchronize_session=False)
    session.query(EntryCatalogItem).filter(
        EntryCatalogItem.id.in_(ec_ids)
    ).delete(synchronize_session=False)
    session.flush()


def _match_structural(row: Dict[str, Any], nodes: List[Dict[str, Any]]) -> Optional[int]:
    serial = _norm(row.get("serial"))
    name = _norm(row.get("name"))
    for node in nodes:
        if not node.get("structural"):
            continue
        if serial and name and serial == _norm(node.get("serial")) and name == _norm(node.get("name")):
            return int(node["id"])
    if name:
        matches = [node for node in nodes if node.get("structural") and name == _norm(node.get("name"))]
        if len(matches) == 1:
            return int(matches[0]["id"])
    return None


def _match_heading_node(row: Dict[str, Any], nodes: List[Dict[str, Any]]) -> Optional[int]:
    serial = _norm(row.get("serial"))
    name = _norm(row.get("name"))
    structural_nodes = [node for node in nodes if node.get("structural")]
    for node in structural_nodes:
        if serial and name and serial == _norm(node.get("serial")) and name == _norm(node.get("name")):
            return int(node["id"])
    if serial:
        matches = [node for node in structural_nodes if serial == _norm(node.get("serial"))]
        if len(matches) == 1:
            return int(matches[0]["id"])
    if name:
        matches = [node for node in structural_nodes if name == _norm(node.get("name"))]
        if len(matches) == 1:
            return int(matches[0]["id"])
    return None


def _looks_structural(row: Dict[str, Any]) -> bool:
    if row.get("pages") or _s(row.get("year")) or _s(row.get("month")) or _s(row.get("day")) or _s(row.get("remark")):
        return False
    name = _s(row.get("name"))
    serial = _s(row.get("serial"))
    return bool(name and serial and name.endswith("材料"))


def _upsert_catalog(session, entry_id: int, template_id: int, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    nodes = _template_nodes(session, template_id)
    used = set()
    current_parent_id: Optional[int] = None
    current_top_heading_id: Optional[int] = None
    targets: List[Dict[str, Any]] = []
    for row in rows:
        if row.get("is_heading") or _is_heading_row(row):
            is_top_heading = _is_chinese_heading_serial(row.get("serial"))
            is_sub_heading = _is_sub_heading_serial(row.get("serial"))
            matched = _match_heading_node(row, nodes)
            if matched is None:
                if is_top_heading:
                    parent_id = None
                elif is_sub_heading:
                    parent_id = current_top_heading_id if current_top_heading_id is not None else current_parent_id
                else:
                    parent_id = current_parent_id
                matched = _create_template_item(session, template_id, parent_id, serial=_s(row.get("serial")), name=_s(row.get("name")))
                nodes = _template_nodes(session, template_id)
            if is_top_heading:
                current_top_heading_id = int(matched)
            current_parent_id = int(matched)
            nodes = _template_nodes(session, template_id)
            continue
        tpl_id = None
        for node in nodes:
            if node.get("structural"):
                continue
            if node.get("parent_id") != current_parent_id:
                continue
            if int(node["id"]) in used:
                continue
            tpl_id = int(node["id"])
            break
        if tpl_id is None:
            tpl_id = _create_template_item(session, template_id, current_parent_id)
            nodes = _template_nodes(session, template_id)
        used.add(int(tpl_id))
        ec = _ensure_ec(session, entry_id, int(tpl_id))
        ec.serial = _s(row.get("serial")) or None
        ec.name = _s(row.get("name")) or None
        ec.year = _s(row.get("year")) or None
        ec.month = _s(row.get("month")) or None
        ec.day = _s(row.get("day")) or None
        ec.pages = int(row["pages"]) if row.get("pages") is not None else None
        ec.remark = _s(row.get("remark")) or None
        session.flush()
        targets.append({"template_item_id": int(tpl_id), "entry_catalog_item_id": int(ec.id), "pages": int(row.get("pages") or 0)})
    if not targets:
        nodes = _template_nodes(session, template_id)
        tpl_id = None
        for node in nodes:
            if not node.get("structural"):
                tpl_id = int(node["id"])
                break
        if tpl_id is None:
            tpl_id = _create_template_item(session, template_id, None)
        ec = _ensure_ec(session, entry_id, tpl_id)
        targets.append({"template_item_id": int(tpl_id), "entry_catalog_item_id": int(ec.id), "pages": 0})
    return targets


def _find_child_dir(folder: str, names: Tuple[str, ...]) -> Optional[str]:
    wanted = {_norm(x) for x in names}
    try:
        children = os.listdir(folder)
    except Exception:
        return None
    for child in children:
        path = os.path.join(folder, child)
        if os.path.isdir(path) and _norm(child) in wanted:
            return path
    return None


def _collect_images(folder: Optional[str]) -> List[str]:
    if not folder or not os.path.exists(folder):
        return []
    out: List[str] = []
    for root, _dirs, files in os.walk(folder):
        for name in files:
            if os.path.splitext(name)[1].lower() in IMAGE_EXTS:
                out.append(os.path.join(root, name))
    out.sort(key=_natural_key)
    return out


def _image_sets(folder: str) -> Tuple[List[str], List[str]]:
    raw_dir = _find_child_dir(folder, ("原图", "原始图", "原始图片", "original", "orig"))
    hd_dir = _find_child_dir(folder, ("高清图", "高清", "hd", "high"))
    processed_dir = _find_child_dir(folder, ("处理图", "处理后", "修图", "retouched", "processed"))
    if raw_dir:
        original_files = _collect_images(raw_dir)
        retouched_files = _collect_images(processed_dir) or _collect_images(hd_dir)
    else:
        original_files = _collect_images(hd_dir)
        retouched_files = _collect_images(processed_dir)
    if not original_files and not retouched_files:
        original_files = _collect_images(folder)
    if not original_files and retouched_files:
        original_files, retouched_files = retouched_files, []
    return original_files, retouched_files


def _copy_file(src: str, dest_dir: str, dest_name: str) -> Tuple[str, int, str]:
    os.makedirs(dest_dir, exist_ok=True)
    dest_name = dest_name if dest_name.lower().endswith(CryptoService.ENCRYPTED_EXT) else dest_name + CryptoService.ENCRYPTED_EXT
    dest_path = os.path.join(dest_dir, dest_name)
    encrypt_image(src, dest_path)
    mime, _ = mimetypes.guess_type(src)
    size = os.path.getsize(dest_path) if os.path.exists(dest_path) else 0
    return dest_path, size, mime or ""


def _upsert_original(session, ec_id: int, file_path: str, file_name: str, file_size: int, mime_type: str, sort_order: int) -> EntryItemImage:
    obj = session.query(EntryItemImage).filter(EntryItemImage.entry_catalog_item_id == int(ec_id), EntryItemImage.image_type == "original", EntryItemImage.file_name == file_name).order_by(EntryItemImage.id.asc()).first()
    if obj is None:
        obj = EntryItemImage(entry_catalog_item_id=int(ec_id), image_type="original", file_name=file_name)
        session.add(obj)
    obj.file_path = file_path
    obj.file_size = file_size
    obj.mime_type = mime_type or ""
    obj.sort_order = int(sort_order)
    obj.original_id = None
    session.flush()
    return obj


def _upsert_retouched(session, original: EntryItemImage, file_path: str, file_name: str, file_size: int, mime_type: str, sort_order: int) -> EntryItemImage:
    obj = session.query(EntryItemImage).filter(EntryItemImage.original_id == int(original.id), EntryItemImage.image_type == "retouched").order_by(EntryItemImage.id.asc()).first()
    if obj is None:
        obj = EntryItemImage(entry_catalog_item_id=int(original.entry_catalog_item_id), image_type="retouched", original_id=int(original.id))
        session.add(obj)
    obj.entry_catalog_item_id = int(original.entry_catalog_item_id)
    obj.file_path = file_path
    obj.file_name = file_name
    obj.file_size = file_size
    obj.mime_type = mime_type or ""
    obj.sort_order = int(sort_order)
    obj.original_id = int(original.id)
    session.flush()
    return obj


def _import_images(session, entry_id: int, image_root: str, targets: List[Dict[str, Any]], original_files: List[str], retouched_files: List[str]) -> Dict[str, Any]:
    imported_original = 0
    imported_retouched = 0
    warnings: List[str] = []
    if not original_files:
        return {"original": 0, "retouched": 0, "warnings": warnings}
    total_pages = sum(int(t.get("pages") or 0) for t in targets)
    if total_pages <= 0:
        targets[0]["pages"] = len(original_files)
        total_pages = len(original_files)
    if len(original_files) != total_pages:
        warnings.append(f"目录页数合计 {total_pages}，高清图 {len(original_files)} 张")
    idx = 0
    for t_index, target in enumerate(targets):
        pages = int(target.get("pages") or 0)
        if pages <= 0:
            continue
        if t_index == len(targets) - 1:
            count = max(pages, len(original_files) - idx)
        else:
            count = pages
        tpl_id = int(target["template_item_id"])
        ec_id = int(target["entry_catalog_item_id"])
        dest_dir = os.path.join(image_root, str(entry_id), str(tpl_id))
        for local_idx in range(count):
            if idx >= len(original_files):
                break
            src = original_files[idx]
            src_name = os.path.basename(src)
            dest_path, size, mime = _copy_file(src, dest_dir, src_name)
            original = _upsert_original(session, ec_id, dest_path, os.path.basename(dest_path), size, mime, local_idx + 1)
            imported_original += 1
            if idx < len(retouched_files):
                rsrc = retouched_files[idx]
                stem, ext = os.path.splitext(os.path.basename(rsrc))
                if not ext:
                    ext = os.path.splitext(src_name)[1]
                rname = f"{os.path.splitext(src_name)[0]}_retouched{ext}"
                rpath, rsize, rmime = _copy_file(rsrc, dest_dir, rname)
                _upsert_retouched(session, original, rpath, os.path.basename(rpath), rsize, rmime, local_idx + 1)
                imported_retouched += 1
            idx += 1
    if len(retouched_files) and len(retouched_files) != imported_retouched:
        warnings.append(f"处理图 {len(retouched_files)} 张，已关联 {imported_retouched} 张")
    return {"original": imported_original, "retouched": imported_retouched, "warnings": warnings}


def _image_root() -> str:
    try:
        root = AppSettings().get_image_root()
        if root:
            return root
    except Exception:
        pass
    return os.getenv("IMAGE_ROOT", os.path.join(os.getcwd(), "data", "images"))


def import_person_archive_folder(folder: str, org_unit_id: Optional[int], template_id: Optional[int] = None) -> Dict[str, Any]:
    folder = os.path.abspath(folder)
    excel_path = _find_excel_file(folder)
    if not excel_path:
        raise ValueError("未找到干部基本信息和档案目录.xlsx")
    tpl_id = int(template_id or get_default_template_id() or 0)
    if tpl_id <= 0:
        raise ValueError("当前系统没有目录模板")
    wb = load_workbook(excel_path, data_only=True)
    basic = _parse_basic_info(wb, folder)
    catalog_rows = _parse_catalog_rows(wb)
    original_files, retouched_files = _image_sets(folder)
    with get_session() as session:
        entry_id, created, match_key = _upsert_entry(session, basic, tpl_id, org_unit_id)
        if catalog_rows and (created or match_key in {"id_card", "emp_no", "name_empty"}):
            _reset_entry_catalog_for_import(session, entry_id)
        targets = _upsert_catalog(session, entry_id, tpl_id, catalog_rows)
        image_result = _import_images(session, entry_id, _image_root(), targets, original_files, retouched_files)
        session.commit()
    return {
        "folder": folder,
        "entry_id": entry_id,
        "name": _s(basic.get("name")),
        "id_card": _s(basic.get("id_card")),
        "created": created,
        "catalog_rows": len(catalog_rows),
        "original_images": image_result.get("original", 0),
        "retouched_images": image_result.get("retouched", 0),
        "warnings": image_result.get("warnings", []),
    }


def discover_person_archive_folders(parent_folder: str) -> List[str]:
    parent_folder = os.path.abspath(parent_folder)
    if _find_excel_file(parent_folder):
        return [parent_folder]
    folders: List[str] = []
    for name in os.listdir(parent_folder):
        path = os.path.join(parent_folder, name)
        if os.path.isdir(path) and _find_excel_file(path):
            folders.append(path)
    folders.sort(key=_natural_key)
    return folders


def import_person_archive_batch(parent_folder: str, org_unit_id: Optional[int], template_id: Optional[int] = None) -> Dict[str, Any]:
    folders = discover_person_archive_folders(parent_folder)
    if not folders:
        raise ValueError("所选目录下未找到可导入的人事档案文件夹")
    results = []
    errors = []
    created = 0
    updated = 0
    original_images = 0
    retouched_images = 0
    for folder in folders:
        try:
            res = import_person_archive_folder(folder, org_unit_id, template_id=template_id)
            results.append(res)
            if res.get("created"):
                created += 1
            else:
                updated += 1
            original_images += int(res.get("original_images") or 0)
            retouched_images += int(res.get("retouched_images") or 0)
        except Exception as e:
            errors.append(f"{os.path.basename(os.path.normpath(folder))}: {e}")
    return {
        "total": len(folders),
        "success": len(results),
        "failed": len(errors),
        "created": created,
        "updated": updated,
        "original_images": original_images,
        "retouched_images": retouched_images,
        "results": results,
        "errors": errors,
    }
