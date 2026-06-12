# -*- coding: utf-8 -*-
"""
导出信息及目录到 Excel（openpyxl）
- Sheet1: 基本信息
- Sheet2: 目录（格式与打印预览一致）
"""

import json
import re
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from common.repositories.entry_repo import (
    get_entry_info,
)
from common.repositories.template_repo import list_catalog_template_items
from ..repo.inventory_entry_repo import get_entry_catalog_item_readonly


# ── 样式常量 ──
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADER_FONT = Font(name="黑体", size=12, bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TITLE_FONT = Font(name="黑体", size=16, bold=True)
_TITLE_ALIGN = Alignment(horizontal="center", vertical="center")

_CATALOG_HEADER_FONT = Font(name="楷体_GB2312", size=11, bold=True)
_CATALOG_DATA_FONT = Font(name="楷体_GB2312", size=11)
_CATALOG_DATA_BOLD = Font(name="楷体_GB2312", size=11, bold=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _s(val) -> str:
    """安全转字符串，None → 空串"""
    if val is None:
        return ""
    return str(val).strip()


def _custom_field_value(info: Dict[str, Any], field_name: str) -> str:
    raw = info.get("custom_fields") or ""
    if not raw:
        return ""
    try:
        fields = json.loads(raw) if isinstance(raw, str) else raw
    except Exception:
        return ""
    if not isinstance(fields, list):
        return ""
    for item in fields:
        if not isinstance(item, dict):
            continue
        if (item.get("field_name") or "").strip() == field_name:
            return _s(item.get("field_value"))
    return ""


def _build_catalog_tree_data(entry_id: int, template_id: int) -> List[Dict[str, Any]]:
    """
    从数据库构建目录树平铺数据（与打印预览格式一致）。
    返回列表，每项: serial, name, year, month, day, pages, remark, is_template
    """
    # 1. 获取模板项（树结构定义）
    tpl_items = list_catalog_template_items(template_id)
    # 按 parent_id 分组
    by_parent: Dict[Optional[int], list] = {}
    for t in tpl_items:
        pk = t.get("parent_id") or None
        by_parent.setdefault(pk, []).append(t)

    result: List[Dict[str, Any]] = []

    def walk(parent_id, depth=0):
        child_infos = []
        has_real_leaf_data = False

        for t in by_parent.get(parent_id, []):
            tpl_serial = _s(t.get("serial"))
            tpl_name = _s(t.get("name"))
            is_structural = bool(tpl_serial or tpl_name)

            ec_item = get_entry_catalog_item_readonly(entry_id=entry_id, template_item_id=int(t["id"]))
            has_data = ec_item.get("id") is not None
            child_infos.append((t, tpl_serial, tpl_name, is_structural, ec_item, has_data))
            if not is_structural and has_data:
                has_real_leaf_data = True

        blank_shown = False

        for t, tpl_serial, tpl_name, is_structural, ec_item, has_data in child_infos:
            if not is_structural and not has_data:
                # 只有在当前父节点没有任何实际录入内容时，才保留一个空行
                if has_real_leaf_data:
                    continue
                if blank_shown:
                    continue
                blank_shown = True
                result.append({
                    "serial": "",
                    "name": "",
                    "year": "",
                    "month": "",
                    "day": "",
                    "pages": "",
                    "remark": "",
                    "is_template": False,
                    "is_blank": True,
                })
            elif is_structural:
                # 模板结构节点（即使没录入也显示）
                result.append({
                    "serial": tpl_serial,
                    "name": tpl_name,
                    "year": "",
                    "month": "",
                    "day": "",
                    "pages": "",
                    "remark": "",
                    "is_template": True,
                    "is_blank": False,
                })
            else:
                result.append({
                    "serial": _s(ec_item.get("serial")) or tpl_serial,
                    "name": _s(ec_item.get("name")) or tpl_name,
                    "year": _s(ec_item.get("year")),
                    "month": _s(ec_item.get("month")),
                    "day": _s(ec_item.get("day")),
                    "pages": _s(ec_item.get("pages")),
                    "remark": _s(ec_item.get("remark")),
                    "is_template": is_structural,
                    "is_blank": False,
                })

            # 递归子节点
            walk(t["id"], depth + 1)

    walk(None, 0)
    return result


def _write_basic_info_sheet(ws, entries_info: List[Dict[str, Any]]):
    """写入"基本信息"工作表"""
    headers = ["序号", "姓名", "工号", "岗位", "电话", "身份证号", "民族", "籍贯", "出生日期", "状态", "所属机构"]
    col_widths = [8, 16, 16, 16, 16, 22, 12, 20, 16, 12, 24]

    # 设置列宽
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="人员基本信息")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _TITLE_ALIGN
    ws.row_dimensions[1].height = 36

    # 表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[2].height = 28

    # 数据行
    for idx, info in enumerate(entries_info, 1):
        row_num = idx + 2
        values = [
            idx,
            info.get("name", ""),
            info.get("emp_no", ""),
            info.get("role_title", ""),
            info.get("phone", ""),
            info.get("id_card", ""),
            _custom_field_value(info, "民族"),
            _custom_field_value(info, "籍贯"),
            _custom_field_value(info, "出生日期"),
            info.get("status", ""),
            info.get("org_unit_name", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER_ALIGN if col != 2 else Alignment(horizontal="left", vertical="center")
            cell.font = Font(name="宋体", size=11)


def _write_catalog_sheet(ws, person_name: str, catalog_data: List[Dict[str, Any]], start_row: int = 1) -> int:
    """
    写入一个人的目录数据到工作表（格式与打印预览一致）。
    返回写完后的下一行行号。
    """
    col_widths = [10, 50, 10, 8, 8, 10, 16]
    for i, w in enumerate(col_widths, 1):
        # 只在首次（start_row==1）设置列宽
        if start_row == 1:
            ws.column_dimensions[get_column_letter(i)].width = w

    r = start_row

    # ── 人员姓名标题 ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    title_cell = ws.cell(row=r, column=1, value=f"干部人事档案目录 — {person_name}")
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _TITLE_ALIGN
    ws.row_dimensions[r].height = 36
    r += 1

    # ── 表头第一行（序号、材料名称跨2行；材料形成时间跨3列；页数、备注跨2行）──
    header_r1 = r
    header_r2 = r + 1

    # 序号 (跨2行)
    ws.merge_cells(start_row=header_r1, start_column=1, end_row=header_r2, end_column=1)
    c = ws.cell(row=header_r1, column=1, value="序号")
    c.font = _CATALOG_HEADER_FONT; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER
    ws.cell(row=header_r2, column=1).border = _THIN_BORDER

    # 材料名称 (跨2行)
    ws.merge_cells(start_row=header_r1, start_column=2, end_row=header_r2, end_column=2)
    c = ws.cell(row=header_r1, column=2, value="材料名称")
    c.font = _CATALOG_HEADER_FONT; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER
    ws.cell(row=header_r2, column=2).border = _THIN_BORDER

    # 材料形成时间 (跨3列)
    ws.merge_cells(start_row=header_r1, start_column=3, end_row=header_r1, end_column=5)
    c = ws.cell(row=header_r1, column=3, value="材料形成时间")
    c.font = _CATALOG_HEADER_FONT; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER
    for cc in range(4, 6):
        ws.cell(row=header_r1, column=cc).border = _THIN_BORDER

    # 年/月/日（第二行）
    for ci, label in zip([3, 4, 5], ["年", "月", "日"]):
        c = ws.cell(row=header_r2, column=ci, value=label)
        c.font = _CATALOG_HEADER_FONT; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER

    # 页数 (跨2行)
    ws.merge_cells(start_row=header_r1, start_column=6, end_row=header_r2, end_column=6)
    c = ws.cell(row=header_r1, column=6, value="页数")
    c.font = _CATALOG_HEADER_FONT; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER
    ws.cell(row=header_r2, column=6).border = _THIN_BORDER

    # 备注 (跨2行)
    ws.merge_cells(start_row=header_r1, start_column=7, end_row=header_r2, end_column=7)
    c = ws.cell(row=header_r1, column=7, value="备注")
    c.font = _CATALOG_HEADER_FONT; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER
    ws.cell(row=header_r2, column=7).border = _THIN_BORDER

    ws.row_dimensions[header_r1].height = 24
    ws.row_dimensions[header_r2].height = 22
    r = header_r2 + 1

    # ── 数据行 ──
    for item in catalog_data:
        is_tpl = item.get("is_template", False)
        font = _CATALOG_DATA_BOLD if is_tpl else _CATALOG_DATA_FONT

        values = [
            item.get("serial", ""),
            item.get("name", ""),
            item.get("year", ""),
            item.get("month", ""),
            item.get("day", ""),
            item.get("pages", ""),
            item.get("remark", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = font
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT_ALIGN if col == 2 else _CENTER_ALIGN
        ws.row_dimensions[r].height = 22
        r += 1

    return r


def export_info_and_catalog(entry_ids: List[int], save_path: str):
    """
    导出选中人员的基本信息和目录到 Excel。
    - Sheet1 "基本信息": 所有人员的基本信息表
    - Sheet2 "目录": 每个人员的目录依次排列（人员之间空一行）
    """
    wb = Workbook()

    # ── Sheet1: 基本信息 ──
    ws_info = wb.active
    ws_info.title = "基本信息"

    entries_info: List[Dict[str, Any]] = []
    for eid in entry_ids:
        info = get_entry_info(entry_id=eid)
        if info:
            entries_info.append(info)

    _write_basic_info_sheet(ws_info, entries_info)

    # ── Sheet2: 目录 ──
    ws_catalog = wb.create_sheet("目录")
    current_row = 1

    for info in entries_info:
        eid = info["id"]
        tpl_id = info.get("template_id")
        person_name = info.get("name", "")

        if tpl_id:
            catalog_data = _build_catalog_tree_data(eid, tpl_id)
        else:
            catalog_data = []

        current_row = _write_catalog_sheet(ws_catalog, person_name, catalog_data, start_row=current_row)
        # 人员之间空一行
        current_row += 1

    wb.save(save_path)


def export_info_and_catalog_batch(
    entry_ids: List[int],
    folder: str,
    id_to_name: Optional[Dict[int, str]] = None,
) -> List[str]:
    """
    批量导出：每个人员生成一个独立 Excel 文件到指定文件夹。
    返回所有生成的文件路径列表。

    参数
    ----
    entry_ids : List[int]
        要导出的人员 entry ID 列表
    folder : str
        目标文件夹路径
    id_to_name : Optional[Dict[int, str]]
        entry_id → 姓名 映射，用于生成文件名；缺失时用 "人员{id}"
    """
    import os
    saved_paths: List[str] = []

    for eid in entry_ids:
        info = get_entry_info(entry_id=eid)
        if not info:
            continue
        person_name = (id_to_name or {}).get(eid) or info.get("name") or f"人员{eid}"
        # 清理文件名中的非法字符
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', person_name)
        filename = f"{safe_name}-信息及目录.xlsx"
        filepath = os.path.join(folder, filename)

        export_info_and_catalog([eid], filepath)
        saved_paths.append(filepath)

    return saved_paths
